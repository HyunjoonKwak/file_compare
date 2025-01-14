import openpyxl
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog, messagebox, Button, Label, StringVar, ttk

def compare_excel_files(file1, file2, output_file, progress_var, root):
    """
    두 Excel 파일을 비교하여 차이를 새 파일에 저장합니다.
    """
    # 파일 불러오기
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # 결과 파일 생성
    wb_result = openpyxl.Workbook()
    wb_result.remove(wb_result.active)  # 기본 시트 제거

    # 색상 정의
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # 추가된 행 (연분홍)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 삭제된 행 (회색)
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 변경된 셀 (파란색)

    total_sheets = len([name for name in wb1.sheetnames if name not in ["Cover", "test description", "Results"]])
    current_sheet = 0

    for sheet_name in wb1.sheetnames:
        if sheet_name in ["Cover", "test description", "Results"]:
            continue

        current_sheet += 1
        progress_var.set(int((current_sheet / total_sheets) * 100))
        root.update_idletasks()  # 진행률 즉시 업데이트

        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        ws_result = wb_result.create_sheet(sheet_name)

        if sheet_name == "History":
            # History 시트 처리 (4열부터 비교)
            rows1 = list(ws1.iter_rows(min_row=2, min_col=4, values_only=True))
            rows2 = list(ws2.iter_rows(min_row=2, min_col=4, values_only=True))

            for i, row in enumerate(rows2):
                if row not in rows1:
                    ws_result.append(["(추가된 행)"] + list(row))

        else:
            # 나머지 시트 처리 (B열 기준)
            rows1 = {row[1]: row for row in ws1.iter_rows(min_row=2, values_only=True) if row[1] and str(row[1]).startswith("IO")}
            rows2 = {row[1]: row for row in ws2.iter_rows(min_row=2, values_only=True) if row[1] and str(row[1]).startswith("IO")}

            keys1 = set(rows1.keys())
            keys2 = set(rows2.keys())

            added_keys = keys2 - keys1
            deleted_keys = keys1 - keys2
            common_keys = keys1 & keys2

            # 추가된 행
            for key in added_keys:
                row = rows2[key]
                ws_result.append(row)
                for cell in ws_result[ws_result.max_row]:
                    cell.fill = pink_fill

            # 삭제된 행
            for key in deleted_keys:
                row = rows1[key]
                ws_result.append(row)
                for cell in ws_result[ws_result.max_row]:
                    cell.fill = gray_fill

            # 변경된 행
            for key in common_keys:
                row1 = rows1[key]
                row2 = rows2[key]
                is_changed = False
                new_row = []
                for cell1, cell2 in zip(row1, row2):
                    if cell1 == cell2:
                        new_row.append(cell2)
                    else:
                        new_row.append(cell2)
                        is_changed = True
                if is_changed:  # 변경된 행만 추가
                    ws_result.append(new_row)
                    for i, (cell1, cell2) in enumerate(zip(row1, row2)):
                        if cell1 != cell2:
                            ws_result.cell(row=ws_result.max_row, column=i + 1).fill = blue_fill

    # 파일 확장자 확인 및 저장
    if not output_file.endswith(".xlsx"):
        output_file += ".xlsx"

    wb_result.save(output_file)
    progress_var.set(100)  # 완료
    root.update_idletasks()  # 마지막 업데이트
    messagebox.showinfo("완료", f"비교 결과가 {output_file}에 저장되었습니다.")

def select_file(label):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        label.config(text=filename)
    return filename

def select_save_file():
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    return filename

def start_comparison():
    file1 = file1_label.cget("text")
    file2 = file2_label.cget("text")
    output_file = select_save_file()

    if not file1 or not file2 or not output_file:
        messagebox.showerror("오류", "모든 파일과 저장 경로를 지정해주세요.")
        return

    try:
        progress_var.set(0)
        compare_excel_files(file1, file2, output_file, progress_var, root)
    except Exception as e:
        messagebox.showerror("오류", f"파일 비교 중 오류가 발생했습니다: {e}")

# Tkinter 인터페이스 생성
root = Tk()
root.title("Excel 파일 비교")

# 진행률 변수 및 Progressbar 생성
progress_var = StringVar()
progress_var.set(0)

Label(root, text="첫 번째 파일:").grid(row=0, column=0, padx=10, pady=5)
file1_label = Label(root, text="", width=50, anchor="w", relief="solid")
file1_label.grid(row=0, column=1, padx=10, pady=5)
Button(root, text="파일 선택", command=lambda: select_file(file1_label)).grid(row=0, column=2, padx=10, pady=5)

Label(root, text="두 번째 파일:").grid(row=1, column=0, padx=10, pady=5)
file2_label = Label(root, text="", width=50, anchor="w", relief="solid")
file2_label.grid(row=1, column=1, padx=10, pady=5)
Button(root, text="파일 선택", command=lambda: select_file(file2_label)).grid(row=1, column=2, padx=10, pady=5)

progress_label = Label(root, text="진행률:")
progress_label.grid(row=2, column=0, padx=10, pady=5)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate", variable=progress_var)
progress_bar.grid(row=2, column=1, columnspan=2, padx=10, pady=5)

# 비교 시작 버튼
Button(root, text="비교 시작 및 저장", command=start_comparison).grid(row=3, column=0, columnspan=3, pady=20)

root.mainloop()
